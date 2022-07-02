import os
from time import sleep

try:
    import openpyxl
    import requests
    import bs4
    from matplotlib import pyplot as plt
    import lxml

except ModuleNotFoundError:
    print('''You must install modules: requests, bs4, openpyxl, matplotlib, lxml.
If You don\'t know how to do this, look on this website:
https://docs.python.org/3/installing/index.html''')
    input()
    quit()


clear = lambda: os.system('cls')


# funkcja znajduje 20 najwazniejszych transferow w podanym roku (od 1970 do 2021)
# zwraca w postaci dictionary ponizsze informacje:
# {name: [(club, club_country), cost, nationality]}
# {imie i nazwisko pilkarza: [(klub wykupujacy, kraj klubu), koszt zawodnika, narodowosc zawodnika]}
def scraping_transfermarkt(year):

    # zrobienie "zupy" ze strony transfermarket
    url = 'https://www.transfermarkt.com/transfers/saisontransfers/statistik/top/plus/0/galerie/0?saison_id={}&transferf' \
          'enster=alle&land_id=&ausrichtung=&spielerposition_id=&altersklasse=&leihe='
    # headers - tego nie rozumiem, ale tak pokazali w necie
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:63.0) Gecko/20100101 Firefox/63.0'}

    result = requests.get(url.format(year), headers=headers)
    soup = bs4.BeautifulSoup(result.text, 'lxml')

    # na stronie transfermarket pilkarze sa podzieleni w wierszach i oznaczeni numerami
    # (pierwszy na stronie - 1,drugi - 2, trzeci - 3, ...), pilkarze o numerach parzystych
    # sa przypisani do innej klasy (w kodzie zrodlowym strony html) niz ci o numerach nieparzytych

    odd_class = soup.select('.odd')  # wyciagniecie danych o pilkarzu (klasa = nieparzysty)
    even_class = soup.select('.even')  # wyciagniecie danych o pilkarzu (klasa = parzystym)

    list_of_footballer = {}
    odd = True

    for i in range(10):
        # petla wyciaga dane pierwszych 20 pilkarzy na stronie
        # name - imie pilkarza,
        # club - klub ktory kupil/wypozyczyl pilkarza
        # cost - kwota wykupu
        # club_country - kraj pochodzenia klubu
        # nationality - narodowosc pilkarza (w przypadku dwoch podaje tylko jedna, pierwsza podana na stronie)

        if odd:
            name = odd_class[i // 2].select('.hauptlink')[0].text
            name = name[1:]
            club = odd_class[i // 2].select('.hauptlink')[1].text
            club = club[1:]
            cost = odd_class[i // 2].select('.hauptlink')[2].text

            club_country = odd_class[i // 2].select('.zentriert')[3].select('.flaggenrahmen')[0]['title']
            nationality = odd_class[i // 2].select('.zentriert')[2].select('.flaggenrahmen')[0]['title']

            odd = False

        elif not odd:
            name = even_class[i // 2].select('.hauptlink')[0].text
            name = name[1:]
            club = even_class[i // 2].select('.hauptlink')[1].text
            club = club[1:]
            cost = even_class[i // 2].select('.hauptlink')[2].text

            club_country = even_class[i // 2].select('.zentriert')[3].select('.flaggenrahmen')[0]['title']
            nationality = even_class[i // 2].select('.zentriert')[2].select('.flaggenrahmen')[0]['title']

            odd = True

        # formatuje tekst przy wypozyczonych pilkarzach
        if 'Loan fee' in cost:
            cost = cost[10:]
            cost = float(cost[:-1])
            club = club + '- loan'

        elif 'transfer' in cost.lower():
            cost = 0.0

        # formatuje wartosc transferu na zmienna float
        else:
            cost = cost[1:]
            cost = float(cost[:-1])

        # dane gromadzone sa w wielopoziomowej zmiennej dict
        list_of_footballer[name] = [(club, club_country), cost, nationality]

    return list_of_footballer


class Footballer:

    def __init__(self, name, club, club_country, cost, nationality):
        self.name = name
        self.club = club
        self.club_country = club_country
        self.cost = cost
        self.nationality = nationality

    def __str__(self):

        if '- loan' in self.club:
            return self.name + 'was loaned by ' + self.club[:-6] + 'for ' + str(self.cost) + ' mln €.'

        else:
            return self.name + 'was bought by ' + self.club + 'for ' + str(self.cost) + ' mln €.'


class ListOfTransfers:

    def __init__(self, year):

        self.year = year
        scraped_list = scraping_transfermarkt(self.year)
        self.footballers = []

        for key, values in scraped_list.items():

            player = Footballer(key, values[0][0], values[0][1], values[1], values[2])
            self.footballers.append(player)

    def sum_cost(self):

        all_value = 0

        for player in self.footballers:
            all_value += player.cost

        return all_value

    def sum_club_countries(self):

        all_club_countries = []

        for player in self.footballers:
            all_club_countries.append(player.club_country)

        club_countries = {country: all_club_countries.count(country) for country in all_club_countries}

        return club_countries

    def sum_nationalities(self):

        all_nationalities = []

        for player in self.footballers:
            all_nationalities.append(player.nationality)

        nationalities = {country: all_nationalities.count(country) for country in all_nationalities}

        return nationalities

    def no_loan(self):

        loans = 0

        for player in self.footballers:

            if '- loan' in player.club:
                loans += 1

        return loans


def create_sheet(wb, players_class):

    sheet = wb.create_sheet(title=f'{players_class.year}')
    sheet['A1'] = 'Piłkarz'
    sheet['B1'] = 'Narodowość'
    sheet['C1'] = 'Klub'
    sheet['D1'] = 'Koszt [mln €]'

    i = 2
    for player in players_class.footballers:
        sheet[f'A{i}'] = player.name
        sheet[f'B{i}'] = player.nationality
        sheet[f'C{i}'] = player.club
        sheet[f'D{i}'] = player.cost
        i += 1

        if player == players_class.footballers[-1]:
            sheet[f'C{13}'] = 'SUMA'
            sheet[f'D{13}'] = players_class.sum_cost()
    return sheet


# data -> dictionary np. {2010: 850,4} -> suma wydanej kasy w danym roku na transfery (10 najwazniejszych)
def create_graph(data1, data2, title1, title2):

    fig, axis = plt.subplots(2, figsize=(7.5, 7.5), num='Best transfers')

    # first graph
    x1 = data1.keys()
    y1 = data1.values()
    graph1 = axis[0].bar(x1, y1, width=0.3)
    axis[0].set_xlabel('Sezon w którym dokonywano transferów')
    axis[0].set_ylabel('Suma cen piłkarzy [mln €]')
    axis[0].bar_label(graph1)
    axis[0].set_title(title1)
    axis[0].set_ylim([0, max(data1.values())+150])

    # second graph
    x2 = data2.keys()
    y2 = data2.values()
    graph2 = axis[1].bar(x2, y2, width=0.6)
    axis[1].set_xlabel('Kraje w których kupowano piłkarzy')
    axis[1].set_ylabel('Liczba kupionych piłkarzy')
    axis[1].bar_label(graph2)
    axis[1].set_title(title2)
    axis[1].set_ylim([0, max(data2.values())+10])

    plt.subplots_adjust(hspace=1)

    plt.show()


# name -> string, years -> tuple: (rok poczatkowy, rok koncowy)
def start_program(name, years):

    wb = openpyxl.Workbook()

    # data -> dictionary np. {2010: 850,4} -> suma wydanej kasy w danym roku na transfery (10 najwazniejszych)
    sum_of_cost = {}

    # data -> dictionary np. {'England': 10} -> ilość transferów w klubach z danego kraju
    sum_club_countries = {}

    for year in range(years[0], years[1]+1):
        best_transfers = ListOfTransfers(year)
        create_sheet(wb, best_transfers)

        # data for graphs -> sum of players cost
        sum_of_cost[str(year)] = best_transfers.sum_cost()

        # data for graphs -> sum of transfers in specific country
        for key, item in best_transfers.sum_club_countries().items():

            if key in sum_club_countries:
                sum_club_countries[key] += item

            else:
                sum_club_countries[key] = item


    del wb['Sheet']
    wb.save(f'{name}.xlsx')
    wb.close()

    file = os.getcwd()
    os.startfile(f'{file}\\{name}.xlsx')
    sleep(2)

    create_graph(sum_of_cost, sum_club_countries, 'Suma cen transferów w danym roku', 'Liczba transferów w danym kraju')


print('''Program służy do wyszukania 10 najważniejszych transferów w wybranym przez użytkownika okresie (1990-2022).
Piłkarze zostaną wypisani w Excelu. Arkusz zostanie automatycznie zapisany. Dodatkowo zostaną wygenerowane dwa wykresy:
   - pierwszy wykres podaje wartości sumaryczne cen piłkarzy w poszczególnych sezonach,
   - drugi wykres sumuje liczbę transferów w poszczególnych krajach.
W programie przyjęto konwencję, że sezon 2015/2016 podaje się jako 2015 itd.
\nZa chwilę zostaniesz poproszony o wpisanie potrzebych informacji.''')
input('Aby przejść dalej naciśnij enter.')
clear()

good_start_year = False
good_end_year = False

while not good_start_year:

    start_year = input('Podaj początek okresu który Cię interesuje i naciśnij enter: ')
    if not start_year.isdecimal():
        clear()
        print("To musi być liczba. Proszę wpisz jeszcze raz.")
    elif int(start_year) < 1990 or int(start_year) > 2022:
        clear()
        print("Proszę o podanie roku z przedziału 1990-2022.")
    elif int(start_year) == 2022:
        end_year = 2022
        clear()
        good_start_year = True
        good_end_year = True
    else:
        good_start_year = True

while not good_end_year:

    end_year = input('Podaj koniec okresu który Cię interesuje i naciśnij enter: ')
    if not end_year.isdecimal():
        clear()
        print("To musi być liczba. Proszę wpisz jeszcze raz.")
    elif int(end_year) < int(start_year) or int(end_year) > 2022:
        clear()
        print(f"Proszę o podanie roku z przedziału 1991-2022, ale większego od początkowego roku ({start_year})")

    else:
        good_end_year = True

clear()
years = (int(start_year), int(end_year))
if years[0] == years[1]:
    print(f'Proszę czekać, trwa wyszukiwanie najważniejszych transferów z sezonu {years[0]}')
else:
    print(f'Proszę czekać, trwa wyszukiwanie najważniejszych transferów z sezonów {years[0]}-{years[1]}')

start_program(f'best_transf_{years[0]}_{years[1]}', years)

clear()
print(f'Plik został zapisany pod ścieżką: {os.getcwd()}')
input('Naciśnij enter, żeby wyjść.')

